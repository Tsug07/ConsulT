import pandas as pd
import unidecode
from thefuzz import fuzz, process
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import customtkinter as ctk
from tkinter import filedialog, messagebox
from PIL import Image
import threading
import os


# === FUNCOES DE LIMPEZA ===
def limpar_cnpj(cnpj):
    if pd.isna(cnpj):
        return ""
    return ''.join([c for c in str(cnpj) if c.isdigit()])


def limpar_nome(nome):
    if pd.isna(nome):
        return ""
    return unidecode.unidecode(str(nome)).strip().upper()


def formatar_cnpj_cpf(digitos):
    # Ate 11 digitos = CPF, acima = CNPJ
    if len(digitos) <= 11:
        digitos = digitos.zfill(11)
        return f"{digitos[:3]}.{digitos[3:6]}.{digitos[6:9]}-{digitos[9:]}"
    else:
        digitos = digitos.zfill(14)
        return f"{digitos[:2]}.{digitos[2:5]}.{digitos[5:8]}/{digitos[8:12]}-{digitos[12:]}"


def nome_parecido_inicio(nome, nomes_ref, limite=90, prefixo=15):
    if not nome:
        return False
    inicio = nome[:prefixo]
    for ref in nomes_ref:
        if fuzz.ratio(inicio, ref[:prefixo]) >= limite:
            return True
    return False


# === LOGICA VERI (antiga) ===
def executar_veri(arquivo_kontrol, arquivo_cac, caminho_saida, log_func, on_done):
    try:
        log_func("Lendo arquivos...")
        e_kontrol = pd.read_excel(arquivo_kontrol, skiprows=2)
        e_cac = pd.read_excel(arquivo_cac, skiprows=2)

        # Selecao de colunas
        # e-Kontrol: C=Nome (indice 2), F=Inscricao Federal/CNPJ (indice 5)
        df_kontrol = e_kontrol.iloc[:, [2, 5]].copy()
        df_kontrol.columns = ['Nome', 'CNPJ']

        # e-CAC: A=Razao Social (indice 0), C=CNPJ (indice 2)
        df_cac = e_cac.iloc[:, [2, 0]].copy()
        df_cac.columns = ['CNPJ', 'Nome']

        # Limpeza
        log_func("Limpando dados...")
        df_kontrol['CNPJ_LIMPO'] = df_kontrol['CNPJ'].apply(limpar_cnpj)
        df_kontrol['NOME_LIMPO'] = df_kontrol['Nome'].apply(limpar_nome)

        df_cac['CNPJ_LIMPO'] = df_cac['CNPJ'].apply(limpar_cnpj)
        df_cac['NOME_LIMPO'] = df_cac['Nome'].apply(limpar_nome)

        # Filtra CNPJs validos
        df_kontrol = df_kontrol[df_kontrol['CNPJ_LIMPO'] != ""]
        df_cac = df_cac[df_cac['CNPJ_LIMPO'] != ""]

        # Etapa 1: comparacao exata de CNPJ
        log_func("Etapa 1: Comparando CNPJs...")
        faltando = df_kontrol[~df_kontrol['CNPJ_LIMPO'].isin(df_cac['CNPJ_LIMPO'])].copy()
        log_func(f"  {len(faltando)} empresas sem CNPJ correspondente.")

        # Etapa 2: fuzzy matching por nome (token_sort_ratio)
        log_func("Etapa 2: Comparando nomes (fuzzy)...")
        nomes_cac = df_cac['NOME_LIMPO'].unique().tolist()

        def nome_existe(nome):
            if not nome:
                return False
            match, score = process.extractOne(nome, nomes_cac, scorer=fuzz.token_sort_ratio)
            return score >= 90

        faltando['Match_Nome'] = faltando['NOME_LIMPO'].apply(nome_existe)
        pendentes = faltando[faltando['Match_Nome'] == False].copy()
        log_func(f"  {len(pendentes)} empresas restantes apos fuzzy por nome completo.")

        # Etapa 3: fuzzy matching por inicio do nome
        log_func("Etapa 3: Verificando inicio dos nomes...")
        pendentes['Existe_Similar'] = pendentes['NOME_LIMPO'].apply(
            lambda x: nome_parecido_inicio(x, nomes_cac)
        )
        resultado = pendentes[pendentes['Existe_Similar'] == False]

        # Salva resultado
        resultado[['Nome', 'CNPJ']].to_excel(caminho_saida, index=False)

        log_func(f"\nConcluido! Arquivo salvo: {caminho_saida}")
        log_func(f"Total de empresas que nao constam: {len(resultado)}")
        on_done(True)

    except Exception as e:
        log_func(f"\nErro: {e}")
        on_done(False)


# === LOGICA CONSUL_ECAC ===
def executar_consul_ecac(arquivo_kontrol, arquivo_comparacao, caminho_saida, tipo_saida, log_func, on_done):
    try:
        log_func("Lendo arquivos...")
        e_kontrol = pd.read_excel(arquivo_kontrol, skiprows=2)
        e_comp = pd.read_excel(arquivo_comparacao, skiprows=2)

        # e-Kontrol: D=Nome (indice 3), G=Inscricao Federal (indice 6), H=Email (indice 7)
        df_kontrol = e_kontrol.iloc[:, [3, 6, 7]].copy()
        df_kontrol.columns = ['Nome', 'CNPJ', 'Email']

        # Consulta e-CAC: A=CNPJ/CPF (indice 0), B=Cliente (indice 1)
        df_comp = e_comp.iloc[:, [0, 1]].copy()
        df_comp.columns = ['CNPJ', 'Nome']

        # Limpeza
        log_func("Limpando dados...")
        df_kontrol['CNPJ_LIMPO'] = df_kontrol['CNPJ'].apply(limpar_cnpj)
        df_kontrol['NOME_LIMPO'] = df_kontrol['Nome'].apply(limpar_nome)

        df_comp['CNPJ_LIMPO'] = df_comp['CNPJ'].apply(limpar_cnpj)
        df_comp['NOME_LIMPO'] = df_comp['Nome'].apply(limpar_nome)

        df_kontrol = df_kontrol[df_kontrol['CNPJ_LIMPO'] != ""]
        df_comp = df_comp[df_comp['CNPJ_LIMPO'] != ""]

        nomes_kontrol = df_kontrol['NOME_LIMPO'].unique().tolist()
        nomes_comp = df_comp['NOME_LIMPO'].unique().tolist()

        def encontra_no_kontrol(cnpj, nome):
            if cnpj in set(df_kontrol['CNPJ_LIMPO']):
                return True
            if not nome:
                return False
            match, score = process.extractOne(nome, nomes_kontrol, scorer=fuzz.token_sort_ratio)
            if score >= 90:
                return True
            return nome_parecido_inicio(nome, nomes_kontrol)

        def encontra_na_comp(cnpj, nome):
            if cnpj in set(df_comp['CNPJ_LIMPO']):
                return True
            if not nome:
                return False
            match, score = process.extractOne(nome, nomes_comp, scorer=fuzz.token_sort_ratio)
            if score >= 90:
                return True
            return nome_parecido_inicio(nome, nomes_comp)

        # Vermelho: esta na comparacao mas NAO no e-Kontrol (saiu da empresa)
        log_func("Verificando empresas que sairam (vermelho)...")
        df_comp['no_kontrol'] = df_comp.apply(
            lambda r: encontra_no_kontrol(r['CNPJ_LIMPO'], r['NOME_LIMPO']), axis=1
        )
        vermelho = df_comp[df_comp['no_kontrol'] == False][['Nome', 'CNPJ']].copy()
        vermelho['Status'] = 'NAO ESTA NO E-KONTROL'
        log_func(f"  {len(vermelho)} empresas marcadas em vermelho.")

        # Verde: esta no e-Kontrol mas NAO na comparacao (precisa adicionar)
        log_func("Verificando empresas a adicionar (verde)...")
        df_kontrol['na_comp'] = df_kontrol.apply(
            lambda r: encontra_na_comp(r['CNPJ_LIMPO'], r['NOME_LIMPO']), axis=1
        )
        verde = df_kontrol[df_kontrol['na_comp'] == False][['Nome', 'CNPJ', 'Email']].copy()
        verde['Status'] = 'ADICIONAR AO SISTEMA'
        log_func(f"  {len(verde)} empresas marcadas em verde.")

        if tipo_saida == "Exportacao":
            # Gera Excel pronto para importar no sistema (somente empresas a adicionar)
            log_func("Gerando Excel de exportacao...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Exportacao"

            ws.append(["CNPJ/CPF", "Contato", "Email"])

            for _, row in verde.iterrows():
                cnpj_fmt = formatar_cnpj_cpf(limpar_cnpj(row['CNPJ']))
                ws.append([cnpj_fmt, row['Nome'], "pessoal@canellaesantos.com.br"])

            wb.save(caminho_saida)
            log_func(f"\nConcluido! Arquivo salvo: {caminho_saida}")
            log_func(f"Total de empresas para exportar: {len(verde)}")
        else:
            # Gera Excel colorido (relatorio)
            log_func("Gerando Excel com cores...")
            fill_vermelho = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
            fill_verde = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

            wb = Workbook()
            ws = wb.active
            ws.title = "Resultado"

            ws.append(["Nome", "CNPJ", "Status"])
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)

            for _, row in vermelho.iterrows():
                ws.append([row['Nome'], row['CNPJ'], row['Status']])
                for cell in ws[ws.max_row]:
                    cell.fill = fill_vermelho

            for _, row in verde.iterrows():
                ws.append([row['Nome'], row['CNPJ'], row['Status']])
                for cell in ws[ws.max_row]:
                    cell.fill = fill_verde

            wb.save(caminho_saida)
            log_func(f"\nConcluido! Arquivo salvo: {caminho_saida}")
            log_func(f"Vermelho (saiu): {len(vermelho)} | Verde (adicionar): {len(verde)}")
        on_done(True)

    except Exception as e:
        log_func(f"\nErro: {e}")
        on_done(False)


# === INTERFACE ===
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("ConsulT - Relatório de Comparação")
        self.geometry("650x620")
        self.resizable(False, False)

        # Icone da janela
        base_dir = os.path.dirname(os.path.abspath(__file__))
        icon_path = os.path.join(base_dir, "assest", "favicon.ico")
        if os.path.exists(icon_path):
            self.iconbitmap(icon_path)

        frame = ctk.CTkFrame(self)
        frame.pack(fill="both", expand=True, padx=15, pady=15)

        # Logo
        logo_path = os.path.join(base_dir, "assest", "ConsulT_logo.png")
        if os.path.exists(logo_path):
            logo_image = ctk.CTkImage(light_image=Image.open(logo_path),
                                      dark_image=Image.open(logo_path),
                                      size=(80, 80))
            logo_label = ctk.CTkLabel(frame, image=logo_image, text="  ConsulT",
                                      font=ctk.CTkFont(size=22, weight="bold"),
                                      compound="left")
            logo_label.pack(pady=(8, 12))

        # Seletor de modo
        ctk.CTkLabel(frame, text="Modo:").pack(anchor="w", padx=10)
        f_modo = ctk.CTkFrame(frame, fg_color="transparent")
        f_modo.pack(fill="x", padx=10, pady=(0, 10))
        self.var_modo = ctk.StringVar(value="Veri")
        ctk.CTkRadioButton(f_modo, text="Veri", variable=self.var_modo,
                           value="Veri", command=self.atualizar_labels).pack(side="left", padx=(0, 15))
        ctk.CTkRadioButton(f_modo, text="Consul_ECAC", variable=self.var_modo,
                           value="Consul_ECAC", command=self.atualizar_labels).pack(side="left")

        # Arquivo e-Kontrol
        ctk.CTkLabel(frame, text="Arquivo e-Kontrol (.xlsx):").pack(anchor="w", padx=10)
        f1 = ctk.CTkFrame(frame, fg_color="transparent")
        f1.pack(fill="x", padx=10, pady=(0, 8))
        self.var_kontrol = ctk.StringVar()
        ctk.CTkEntry(f1, textvariable=self.var_kontrol).pack(side="left", fill="x", expand=True)
        ctk.CTkButton(f1, text="Buscar", width=80, command=lambda: self.escolher_arquivo(self.var_kontrol)).pack(side="right", padx=(5, 0))

        # Arquivo de comparacao (e-CAC ou outro)
        self.label_comparacao = ctk.CTkLabel(frame, text="Arquivo de comparacao (.xlsx):")
        self.label_comparacao.pack(anchor="w", padx=10)
        f2 = ctk.CTkFrame(frame, fg_color="transparent")
        f2.pack(fill="x", padx=10, pady=(0, 8))
        self.var_comparacao = ctk.StringVar()
        ctk.CTkEntry(f2, textvariable=self.var_comparacao).pack(side="left", fill="x", expand=True)
        ctk.CTkButton(f2, text="Buscar", width=80, command=lambda: self.escolher_arquivo(self.var_comparacao)).pack(side="right", padx=(5, 0))

        # Tipo de saida (so aparece no modo Consul_ECAC)
        self.frame_tipo_saida = ctk.CTkFrame(frame, fg_color="transparent")
        ctk.CTkLabel(self.frame_tipo_saida, text="Tipo de saida:").pack(anchor="w")
        f_tipo = ctk.CTkFrame(self.frame_tipo_saida, fg_color="transparent")
        f_tipo.pack(fill="x", pady=(0, 4))
        self.var_tipo_saida = ctk.StringVar(value="Relatorio")
        ctk.CTkRadioButton(f_tipo, text="Relatorio (com cores)", variable=self.var_tipo_saida,
                           value="Relatorio").pack(side="left", padx=(0, 15))
        ctk.CTkRadioButton(f_tipo, text="Exportacao (CNPJ, Contato, Email)", variable=self.var_tipo_saida,
                           value="Exportacao").pack(side="left")

        # Arquivo de saida
        self.frame_saida = ctk.CTkFrame(frame, fg_color="transparent")
        self.frame_saida.pack(fill="x", padx=10, pady=(0, 12))
        ctk.CTkLabel(self.frame_saida, text="Arquivo de saida (.xlsx):").pack(anchor="w")
        f3 = ctk.CTkFrame(self.frame_saida, fg_color="transparent")
        f3.pack(fill="x")
        self.var_saida = ctk.StringVar(value="Nao_consta_final.xlsx")
        ctk.CTkEntry(f3, textvariable=self.var_saida).pack(side="left", fill="x", expand=True)
        ctk.CTkButton(f3, text="Buscar", width=80, command=self.escolher_saida).pack(side="right", padx=(5, 0))

        # Botao executar
        self.btn_exec = ctk.CTkButton(frame, text="Executar", command=self.iniciar)
        self.btn_exec.pack(pady=(0, 8))

        # Log
        ctk.CTkLabel(frame, text="Log:").pack(anchor="w", padx=10)
        self.log = ctk.CTkTextbox(frame, height=200, state="disabled")
        self.log.pack(fill="both", expand=True, padx=10, pady=(0, 10))

    def atualizar_labels(self):
        modo = self.var_modo.get()
        if modo == "Veri":
            self.label_comparacao.configure(text="Arquivo e-CAC (.xlsx):")
            self.title("Consul - Veri")
            self.frame_tipo_saida.pack_forget()
        else:
            self.label_comparacao.configure(text="Arquivo de comparacao (.xlsx):")
            self.title("Consul - Consul_ECAC")
            self.frame_tipo_saida.pack(fill="x", padx=10, pady=(0, 8), before=self.frame_saida)

    def escolher_arquivo(self, var):
        caminho = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if caminho:
            var.set(caminho)

    def escolher_saida(self):
        caminho = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if caminho:
            self.var_saida.set(caminho)

    def adicionar_log(self, texto):
        self.log.configure(state="normal")
        self.log.insert("end", texto + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    def log_thread_safe(self, texto):
        self.after(0, self.adicionar_log, texto)

    def ao_finalizar(self, sucesso):
        def _f():
            self.btn_exec.configure(state="normal")
            if sucesso:
                messagebox.showinfo("Concluido", "Processamento finalizado com sucesso!")
        self.after(0, _f)

    def iniciar(self):
        kontrol = self.var_kontrol.get().strip()
        comparacao = self.var_comparacao.get().strip()
        saida = self.var_saida.get().strip()
        modo = self.var_modo.get()

        if not kontrol or not comparacao:
            messagebox.showwarning("Atencao", "Selecione os dois arquivos de entrada.")
            return
        if not saida:
            messagebox.showwarning("Atencao", "Defina o arquivo de saida.")
            return

        self.log.configure(state="normal")
        self.log.delete("1.0", "end")
        self.log.configure(state="disabled")

        self.btn_exec.configure(state="disabled")

        if modo == "Veri":
            args = (kontrol, comparacao, saida, self.log_thread_safe, self.ao_finalizar)
            target = executar_veri
        else:
            tipo = self.var_tipo_saida.get()
            args = (kontrol, comparacao, saida, tipo, self.log_thread_safe, self.ao_finalizar)
            target = executar_consul_ecac

        t = threading.Thread(target=target, args=args, daemon=True)
        t.start()


if __name__ == "__main__":
    app = App()
    app.mainloop()
